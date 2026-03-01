#!/usr/bin/env python3
import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

import requests
try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]


DEFAULT_QUERY_HASH = "37479f2b8209594dde7facb0d904896a"
DEFAULT_APP_ID = "936619743392459"
DEFAULT_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

FOLLOWER_FIELDS = [
    "id",
    "username",
    "instagram_url",
    "full_name",
    "is_private",
    "is_verified",
    "profile_pic_url",
    "followed_by_viewer",
    "requested_by_viewer",
    "biography",
    "category",
]


def save_workbook_or_fail(workbook: Workbook, excel_path: Path) -> None:
    try:
        workbook.save(excel_path)
    except OSError as exc:
        raise RuntimeError(f"Failed saving Excel {excel_path}: {exc}") from exc


def get_or_create_sheet_with_header(workbook: Workbook) -> Any:
    if "users" in workbook.sheetnames:
        sheet = workbook["users"]
    else:
        sheet = workbook.active
        sheet.title = "users"

    if sheet.max_row == 0:
        sheet.append(FOLLOWER_FIELDS)
    else:
        header = [sheet.cell(row=1, column=i + 1).value for i in range(len(FOLLOWER_FIELDS))]
        if header != FOLLOWER_FIELDS:
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(FOLLOWER_FIELDS)
    return sheet


def load_seen_ids_from_sheet(sheet: Any) -> set[str]:
    seen_ids: set[str] = set()
    if sheet.max_row < 2:
        return seen_ids

    id_column = 1
    for row_idx in range(2, sheet.max_row + 1):
        follower_id = sheet.cell(row=row_idx, column=id_column).value
        if follower_id is None:
            continue
        seen_ids.add(str(follower_id))
    return seen_ids


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch Instagram users by user_id and export them to Excel (.xlsx)."
    )
    parser.add_argument("--user-id", required=True, help="Instagram user id")
    parser.add_argument("--first", type=int, default=50, help="GraphQL page size")
    parser.add_argument("--after", default="", help="GraphQL pagination cursor")
    parser.add_argument("--query-hash", default=DEFAULT_QUERY_HASH, help="GraphQL query_hash")
    parser.add_argument(
        "--edge-field",
        choices=["auto", "edge_followed_by", "edge_follow"],
        default="auto",
        help="GraphQL edge field to parse: followers(edge_followed_by) or following(edge_follow)",
    )
    parser.add_argument("--cookie", help="Raw cookie header value")
    parser.add_argument("--x-ig-app-id", default=DEFAULT_APP_ID, help="Value for x-ig-app-id header")
    parser.add_argument("--x-asbd-id", help="Value for x-asbd-id header")
    parser.add_argument("--user-agent", default=DEFAULT_UA, help="User-Agent header value")
    parser.add_argument("--timeout", type=float, default=20, help="Request timeout in seconds")
    parser.add_argument("--sleep", type=float, default=0, help="Sleep seconds between follower detail requests")
    parser.add_argument("--excel-path", help="Output Excel path (default: followers_<user_id>.xlsx)")
    return parser.parse_args()


def build_session(args: argparse.Namespace) -> requests.Session:
    session = requests.Session()
    headers = {
        "Accept": "application/json",
        "User-Agent": args.user_agent,
        "x-ig-app-id": args.x_ig_app_id,
    }
    if args.x_asbd_id:
        headers["x-asbd-id"] = args.x_asbd_id
    if args.cookie:
        headers["Cookie"] = args.cookie
    session.headers.update(headers)
    return session


def safe_json(response: requests.Response) -> Tuple[Optional[Any], Optional[str]]:
    try:
        return response.json(), None
    except ValueError:
        text = response.text or ""
        return None, text[:500]


def make_get_request(session: requests.Session, url: str, timeout: float) -> Tuple[Optional[requests.Response], Optional[str]]:
    try:
        response = session.get(url, timeout=timeout)
        return response, None
    except requests.RequestException as exc:
        return None, str(exc)


def print_section(name: str, url: str, status_code: Optional[int], payload: Dict[str, Any]) -> None:
    print(
        json.dumps(
            {"section": name, "url": url, "status_code": status_code, "data": payload},
            ensure_ascii=False
        )
    )


def auth_hint(status_code: int) -> Optional[str]:
    if status_code in (401, 403):
        return "Authentication may be required or expired. Try providing a valid --cookie."
    if status_code == 429:
        return "Rate-limited by Instagram. Try again later or rotate session/cookie."
    return None


def fail(msg: str, code: int = 1) -> int:
    print(json.dumps({"error": msg}, ensure_ascii=False, indent=2), file=sys.stderr)
    return code


def build_graphql_url(query_hash: str, user_id: str, after: str, first: int) -> str:
    variables = {"id": user_id, "after": after, "first": first}
    encoded_variables = quote(json.dumps(variables, separators=(",", ":")))
    return f"https://www.instagram.com/graphql/query/?query_hash={quote(query_hash)}&variables={encoded_variables}"


def get_nested(data: Any, path: List[str]) -> Any:
    node = data
    for key in path:
        if not isinstance(node, dict) or key not in node:
            return None
        node = node[key]
    return node


def resolve_edge_field(user_payload: Dict[str, Any], preferred: str) -> Optional[str]:
    if preferred in ("edge_followed_by", "edge_follow"):
        if isinstance(user_payload.get(preferred), dict):
            return preferred
        return None

    for candidate in ("edge_follow", "edge_followed_by"):
        if isinstance(user_payload.get(candidate), dict):
            return candidate
    return None


def extract_user_page(
    payload: Any, preferred_edge_field: str
) -> Tuple[Optional[int], List[Dict[str, Any]], bool, str, Optional[str]]:
    user_payload = get_nested(payload, ["data", "user"])
    if not isinstance(user_payload, dict):
        return None, [], False, "", None

    resolved_edge_field = resolve_edge_field(user_payload, preferred_edge_field)
    if not resolved_edge_field:
        return None, [], False, "", None

    edge_root = user_payload.get(resolved_edge_field)
    if not isinstance(edge_root, dict):
        return None, [], False, "", None

    total_count = edge_root.get("count")
    edges = edge_root.get("edges") if isinstance(edge_root.get("edges"), list) else []
    page_info = edge_root.get("page_info") if isinstance(edge_root.get("page_info"), dict) else {}

    has_next_page = bool(page_info.get("has_next_page"))
    end_cursor = str(page_info.get("end_cursor") or "")

    followers: List[Dict[str, Any]] = []
    for edge in edges:
        if not isinstance(edge, dict):
            continue
        node = edge.get("node")
        if not isinstance(node, dict):
            continue

        row = {
            "id": node.get("id"),
            "username": node.get("username"),
            "instagram_url": "",
            "full_name": node.get("full_name"),
            "is_private": node.get("is_private"),
            "is_verified": node.get("is_verified"),
            "profile_pic_url": node.get("profile_pic_url"),
            "followed_by_viewer": node.get("followed_by_viewer"),
            "requested_by_viewer": node.get("requested_by_viewer"),
            "biography": "",
            "category": "",
        }
        if row["id"] is None:
            continue
        username = row.get("username")
        if isinstance(username, str) and username:
            row["instagram_url"] = f"https://www.instagram.com/{username}"
        followers.append(row)

    return total_count, followers, has_next_page, end_cursor, resolved_edge_field


def extract_biography_and_category(payload: Any) -> Tuple[str, str]:
    user = get_nested(payload, ["user"])
    if not isinstance(user, dict):
        user = get_nested(payload, ["data", "user"])
        if not isinstance(user, dict):
            return "", ""

    biography = user.get("biography")
    if not isinstance(biography, str):
        biography = ""

    category = user.get("category")
    if not isinstance(category, str) or not category:
        category = user.get("category_name")
    if not isinstance(category, str) or not category:
        category = user.get("business_category_name")
    if not isinstance(category, str):
        category = ""

    return biography, category


def enrich_row_with_detail(
    session: requests.Session,
    row: Dict[str, Any],
    timeout: float,
    query_user_id: str,
) -> bool:
    follower_id = str(row.get("id", ""))
    if not follower_id:
        return False

    detail_url = f"https://i.instagram.com/api/v1/users/{quote(follower_id)}/info/"
    resp, err = make_get_request(session, detail_url, timeout)
    print(
        json.dumps(
            {
                "section": "detail_request",
                "query_user_id": query_user_id,
                "follower_id": follower_id,
                "url": detail_url,
                "status_code": resp.status_code if resp is not None else None,
                "error": err,
            },
            ensure_ascii=False,
        )
    )
    if err:
        return False
    assert resp is not None

    if resp.status_code >= 400:
        return False

    payload, _ = safe_json(resp)
    if payload is None:
        return False

    biography, category = extract_biography_and_category(payload)
    row["biography"] = biography
    row["category"] = category
    return True


def main() -> int:
    args = parse_args()
    session = build_session(args)
    user_id = str(args.user_id)
    relation_hint = "followers" if args.edge_field != "edge_follow" else "following"
    excel_path = Path(args.excel_path) if args.excel_path else Path(f"{relation_hint}_{user_id}.xlsx")

    if Workbook is None or load_workbook is None:
        return fail("Missing dependency 'openpyxl'. Install it with: pip install openpyxl")

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        try:
            workbook = load_workbook(excel_path)
        except OSError as exc:
            return fail(f"Failed opening Excel {excel_path}: {exc}")
    else:
        workbook = Workbook()

    sheet = get_or_create_sheet_with_header(workbook)
    seen_ids = load_seen_ids_from_sheet(sheet)

    cursor = args.after
    page = 1
    total_count: Optional[int] = None
    resolved_edge_field: Optional[str] = None
    count_exported = 0
    enriched_ok = 0
    enriched_fail = 0

    try:
        while True:
            graphql_url = build_graphql_url(args.query_hash, user_id, cursor, args.first)
            resp, err = make_get_request(session, graphql_url, args.timeout)
            if err:
                return fail(f"GraphQL request failed on page {page}: {err}")

            assert resp is not None
            payload, nonjson = safe_json(resp)

            if payload is not None:
                print_section(
                    "graphql_basic_info" if page == 1 else f"graphql_page_{page}",
                    graphql_url,
                    resp.status_code,
                    payload,
                )
            else:
                print_section(
                    "graphql_basic_info" if page == 1 else f"graphql_page_{page}",
                    graphql_url,
                    resp.status_code,
                    {"parse_error": "Non-JSON response", "response_snippet": nonjson},
                )

            hint = auth_hint(resp.status_code)
            if hint:
                print(json.dumps({"warning": hint}, ensure_ascii=False, indent=2), file=sys.stderr)

            if resp.status_code >= 400:
                return fail(f"GraphQL request returned HTTP {resp.status_code} on page {page}.")
            if payload is None:
                return fail(f"GraphQL response is not JSON on page {page}.")

            preferred_edge_field = resolved_edge_field or args.edge_field
            page_total_count, followers, has_next_page, end_cursor, detected_edge_field = extract_user_page(
                payload, preferred_edge_field
            )
            if not detected_edge_field:
                return fail(
                    "Could not find edge_follow/edge_followed_by in response. "
                    "Check --query-hash and --edge-field."
                )
            resolved_edge_field = detected_edge_field

            if page_total_count is not None:
                total_count = page_total_count

            for follower in followers:
                follower_id = str(follower.get("id"))
                if follower_id in seen_ids:
                    continue
                seen_ids.add(follower_id)

                if enrich_row_with_detail(session, follower, args.timeout, user_id):
                    enriched_ok += 1
                else:
                    enriched_fail += 1

                sheet.append([follower.get(field) for field in FOLLOWER_FIELDS])
                count_exported += 1
                save_workbook_or_fail(workbook, excel_path)

                if args.sleep > 0:
                    time.sleep(args.sleep)

            if not has_next_page:
                break
            if not end_cursor:
                return fail("GraphQL indicates next page, but end_cursor is missing.")

            cursor = end_cursor
            page += 1
            print(f'=========>>>>>>>>>user_id:{user_id} page: {page}')
    except RuntimeError as exc:
        return fail(str(exc))
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    print(
        json.dumps(
            {
                "summary": {
                    "user_id": user_id,
                    "total_count_reported": total_count,
                    "edge_field": resolved_edge_field,
                    "rows_already_in_excel": len(seen_ids) - count_exported,
                    "rows_exported_this_run": count_exported,
                    "detail_ok": enriched_ok,
                    "detail_fail": enriched_fail,
                    "excel_path": str(excel_path),
                }
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
